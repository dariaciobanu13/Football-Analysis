import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl.styles import PatternFill
import os
import math
import numpy as np

def generate_incident_log(incidents_df, filename="Performance_Alerts_Log.xlsx"):
    """
    Exports the critical underperformance alerts to an Excel spreadsheet and applies visual alerts (Red formatting).
    """
    print(f"[INFO] Generating Automated Performance Alerts Report: {filename}...")
    
    # Select columns for the final executive report (including Enterprise ROI)
    report_columns = [
        'Time_Window', 'Team_Name', 'Player_ID', 'Player_Name', 
        'Total_Actions', 'Performance_Rolling_15m', 'Performance_Z_Score', 
        'Market_Value_M€', 'ROI_Efficiency', 'Player_Status'
    ]
    
    report_data = incidents_df[report_columns].copy()
    
    # Sort by time so the coach can see a timeline of issues
    report_data = report_data.sort_values(by=['Time_Window', 'Performance_Z_Score'])
    
    # Write to Excel
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        report_data.to_excel(writer, index=False, sheet_name='Performance_Alerts')
        
        # Access the openpyxl workbook and sheet to add colors (Data Engineering touch)
        workbook  = writer.book
        worksheet = writer.sheets['Performance_Alerts']
        
        # Define a red fill for Critical Alerts
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Iterate through the rows and color those labeled 'FATIGUE ALERT (CRITICAL)'
        # In report_columns, 'Player_Status' is index 9 (0-indexed), which is column 10 (1-indexed) in Excel
        for row in range(2, len(report_data) + 2):  # Skip header (row 1)
            status_val = worksheet.cell(row=row, column=10).value
            if status_val == 'FATIGUE ALERT (CRITICAL)':
                # Color the whole row red
                for col in range(1, 11):
                    worksheet.cell(row=row, column=col).fill = red_fill

    print(f"[INFO] Performance Alerts Log successfully saved to {os.path.abspath(filename)}")

def plot_degradation_timeline(team_df, team_name, filename="Performance_Timeline.png"):
    """
    Visualizes Performance Drop-off over time.
    Plots the Performance Index of players to show when they fatigued or lost form.
    """
    print(f"[INFO] Rendering Performance Timeline Plot for team: {team_name}...")
    
    # Filter data for the requested team
    df = team_df[team_df['Team_Name'] == team_name].copy()
    
    plt.figure(figsize=(14, 8))
    sns.set_style("whitegrid") # Cleaner look for general audience
    
    # Plot a line for each primary player (player with >= 20 actions in the match)
    action_counts = df.groupby('Player_Name')['Total_Actions'].sum()
    primary_players = action_counts[action_counts >= 20].index
    
    plot_data = df[df['Player_Name'].isin(primary_players)]
    
    # Map colors for easier distinction
    palette = sns.color_palette("husl", len(primary_players))
    
    # Create the lineplot
    ax = sns.lineplot(
        data=plot_data,
        x='Time_Window', 
        y='Performance_Rolling_15m',
        hue='Player_Name', 
        palette=palette,
        marker='o',
        linewidth=2.5,
        alpha=0.85
    )
    
    # Add Shading for the "Danger Zone" (Underperformance)
    plt.axhspan(ymin=0, ymax=80, color='red', alpha=0.1, label='Underperformance Zone (< 80%)')
    
    # Add a horizontal line representing the Absolute Threshold
    plt.axhline(y=80, color='darkred', linestyle='--', linewidth=1.5)
    
    # Styling the plot
    plt.title(f"Match Timeline: Player Performance & Fatigue - {team_name}", fontsize=18, fontweight='bold', pad=15)
    plt.xlabel("Match Minute", fontsize=14, fontweight='bold')
    plt.ylabel("Performance Score (15m Rolling Avg %)", fontsize=14, fontweight='bold')
    plt.ylim(0, 105) # Score is 0-100%
    plt.xlim(0, 95) # Standard match time + stoppage
    
    # Move legend outside the plot
    plt.legend(bbox_to_anchor=(1.02, 1), loc='upper left', borderaxespad=0., title='Players', fontsize=11, title_fontsize=12)
    
    plt.tight_layout()
    plt.savefig(filename, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"[INFO] Performance Timeline plotted and saved to {os.path.abspath(filename)}")

def plot_node_stability(team_df, team_name, filename="Consistency_Violin.png"):
    """
    Renders a Violin Plot to analyze the consistency (variance) of a player's performance.
    A thin, flat structure means consistent form. A wide, bulging shape means highly volatile performance.
    """
    print(f"[INFO] Rendering Consistency Distribution (Violin Plot) for team: {team_name}...")
    
    df = team_df[team_df['Team_Name'] == team_name].copy()
    
    if df.empty:
        print("[WARNING] No data available for Consistency Plot.")
        return
        
    plt.figure(figsize=(16, 10))
    sns.set_style("whitegrid")
    
    # Create the violin plot
    ax = sns.violinplot(
        data=df,
        x='QoS_Index', 
        y='Player_Name',
        inner='quartile', # Shows median and quartiles inside the violin
        palette="viridis",
        orient='h',
        cut=0, # Prevents the violin from drawing impossible data (< 0 or > 100)
        density_norm='width' # Equal width for better visual comparison
    )
    
    plt.title(f"Player Consistency Distribution - {team_name}", fontsize=18, fontweight='bold', pad=15)
    plt.xlabel("Performance Score (%)", fontsize=14, fontweight='bold')
    plt.ylabel("Players", fontsize=14, fontweight='bold')
    
    # Threshold line
    plt.axvline(x=80, color='red', linestyle='--', linewidth=2, label='Acceptable Form Threshold (80%)')
    
    # Add an explanatory text box (Legend)
    explanation = (
        "How to read this plot:\n"
        "- Wide Bulge: Player spent most of the match at this performance level.\n"
        "- Long/Stretched: Player was highly inconsistent.\n"
        "- Compact/Thin: Player maintained a steady form."
    )
    plt.figtext(0.99, 0.02, explanation, horizontalalignment='right', 
                fontsize=11, bbox=dict(facecolor='white', alpha=0.8, edgecolor='gray', boxstyle='round,pad=0.5'))
    
    plt.legend(loc='lower right')
    
    plt.tight_layout()
    plt.savefig(filename, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"[INFO] Consistency Plot saved to {os.path.abspath(filename)}")

def plot_improvement_radar(team_df, team_name, target_players, filename="Improvement_Radar.png"):
    """
    Creates a Radar (Spider) Chart comparing the targeted underperforming players
    against the Team Average across key metrics.
    """
    print(f"[INFO] Rendering Improvement Radar Chart for isolated players...")
    
    # Filter for the team
    df = team_df[team_df['Team_Name'] == team_name].copy()
    
    # Calculate overall metrics per player
    # Since team_df is derived from agg_df, it has Pass_Acc, Duel_Acc, QoS_Index overall
    player_stats = df.groupby('Player_Name').agg(
        Pass_Accuracy=('Pass_Acc', 'mean'),
        Duel_Win_Rate=('Duel_Acc', 'mean'),
        Overall_Performance=('QoS_Index', 'mean')
    ).reset_index()
    
    # Convert passing and duels from decimal to percentage
    player_stats['Pass_Accuracy'] *= 100
    player_stats['Duel_Win_Rate'] *= 100
    
    # Fill any NaN with 0
    player_stats = player_stats.fillna(0)
    
    # Overall Team Average for comparison
    team_avg = {
        'Pass_Accuracy': player_stats['Pass_Accuracy'].mean(),
        'Duel_Win_Rate': player_stats['Duel_Win_Rate'].mean(),
        'Overall_Performance': player_stats['Overall_Performance'].mean()
    }
    
    labels = ['Pass Accuracy', 'Duel Win Rate', 'Overall Perf']
    num_vars = len(labels)
    
    # Compute angle for each axis
    angles = [n / float(num_vars) * 2 * math.pi for n in range(num_vars)]
    angles += angles[:1] # Close the circle
    
    plt.figure(figsize=(10, 8))
    # Initialize the spider plot
    ax = plt.subplot(111, polar=True)
    
    # Draw one axe per variable and add labels
    plt.xticks(angles[:-1], labels, color='black', size=12, fontweight='bold')
    
    # Draw ylabels
    ax.set_rlabel_position(0)
    plt.yticks([20, 40, 60, 80], ["20", "40", "60", "80"], color="grey", size=10)
    plt.ylim(0, 100)
    
    # Plot Team Average
    team_values = [team_avg['Pass_Accuracy'], team_avg['Duel_Win_Rate'], team_avg['Overall_Performance']]
    team_values += team_values[:1]
    ax.plot(angles, team_values, linewidth=2, linestyle='dashed', label=f"{team_name} Average", color='black')
    ax.fill(angles, team_values, alpha=0.1, color='gray')
    
    # Plot Target Players
    colors = ['red', 'orange', 'purple']
    for i, player in enumerate(target_players):
        player_data = player_stats[player_stats['Player_Name'] == player]
        if not player_data.empty:
            p_vals = player_data.iloc[0]
            values = [p_vals['Pass_Accuracy'], p_vals['Duel_Win_Rate'], p_vals['Overall_Performance']]
            values += values[:1]
            ax.plot(angles, values, linewidth=2, linestyle='solid', label=player, color=colors[i % len(colors)])
            ax.fill(angles, values, alpha=0.15, color=colors[i % len(colors)])
            
    plt.title(f"Performance Gap Analysis (Need Improvement vs {team_name} Avg)", size=16, fontweight='bold', y=1.1)
    plt.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1))
    
    plt.tight_layout()
    plt.savefig(filename, dpi=300)
    plt.close()
    print(f"[INFO] Radar Chart saved to {os.path.abspath(filename)}")

def plot_weather_impact(agg_df, weather_severity, team_name, filename="Weather_Impact_Analysis.png"):
    """
    Visualizes the potential performance loss due to environmental stress.
    """
    print(f"[INFO] Rendering Weather Impact Analysis for {team_name}...")
    
    # Calculate what the performance would be vs what it is
    df = agg_df[agg_df['Team_Name'] == team_name].copy()
    
    # Estimate the "Raw" Performance without weather penalty
    # (Inverse of the penalty we applied in calculations)
    df['Potential_Performance'] = df['QoS_Index'] / (1 - (weather_severity * 0.5 * 0.6))
    
    player_stats = df.groupby('Player_Name').agg(
        Observed=('QoS_Index', 'mean'),
        Potential=('Potential_Performance', 'mean')
    ).reset_index().sort_values(by='Observed', ascending=False)
    
    plt.figure(figsize=(12, 8))
    sns.set_style("whitegrid")
    
    # Plotting
    sns.barplot(data=player_stats, x='Potential', y='Player_Name', color='gray', alpha=0.3, label='Potential (Clear Skies)')
    sns.barplot(data=player_stats, x='Observed', y='Player_Name', color='skyblue', label=f'Observed (Under {weather_severity*100}% Weather Stress)')
    
    plt.title(f"Environmental Impact Report: Weather Degradation - {team_name}", fontsize=16, fontweight='bold')
    plt.xlabel("Performance Score (%)", fontsize=12)
    plt.ylabel("Players", fontsize=12)
    plt.legend(loc='lower right')
    
    plt.tight_layout()
    plt.savefig(filename, dpi=300)
    plt.close()
    print(f"[INFO] Weather Impact Plot saved to {os.path.abspath(filename)}")

def plot_roi_efficiency(agg_df, team_name, filename="Financial_ROI_Matrix.png"):
    """
    Visualizes the Return on Investment (ROI) efficiency of players.
    Lower efficiency score = Better value for money.
    """
    print(f"[INFO] Rendering Financial ROI Efficiency Matrix for {team_name}...")
    
    df = agg_df[agg_df['Team_Name'] == team_name].copy()
    
    player_summary = df.groupby('Player_Name').agg(
        Market_Value=('Market_Value_M€', 'first'),
        ROI=('ROI_Efficiency', 'mean'),
        Avg_Perf=('QoS_Index', 'mean')
    ).reset_index().sort_values(by='ROI', ascending=True)
    
    plt.figure(figsize=(14, 8))
    
    # Bubble chart: Market Value (Size), Performance (X), ROI (Color)
    scatter = plt.scatter(
        x=player_summary['Avg_Perf'], 
        y=player_summary['ROI'], 
        s=player_summary['Market_Value'] * 20, # Scale for visibility
        c=player_summary['ROI'], 
        cmap='RdYlGn_r', # Red to Green (Reversed because lower ROI is better)
        alpha=0.6,
        edgecolors="black"
    )
    
    for i, txt in enumerate(player_summary['Player_Name']):
        plt.annotate(txt, (player_summary['Avg_Perf'].iloc[i], player_summary['ROI'].iloc[i]), fontsize=9)
    
    plt.colorbar(scatter, label='ROI Efficiency (M€ per 1% Performance)')
    plt.title(f"Enterprise ROI Intelligence: Performance vs Financial Capital - {team_name}", fontsize=16, fontweight='bold')
    plt.xlabel("Average Performance Score (%)", fontsize=12)
    plt.ylabel("ROI Index (Lower is More Efficient)", fontsize=12)
    
    plt.tight_layout()
    plt.savefig(filename, dpi=300)
    plt.close()
    print(f"[INFO] Financial ROI Plot saved to {os.path.abspath(filename)}")

def plot_kafka_stream_performance(events_count, duration, filename="Kafka_Stream_Flow.png"):
    """
    Visualizes the throughput of the real-time simulation.
    """
    print(f"[INFO] Generating Kafka Stream Flow Visualization...")
    
    labels = ['Real-time Streamed', 'Batch Processed']
    values = [events_count, events_count] # Just to compare architectures
    
    plt.figure(figsize=(8, 6))
    colors = ['#FF9900', '#0073BB'] # Kafka vs AWS/Batch colors
    
    bars = plt.bar(labels, values, color=colors, alpha=0.8)
    plt.title("System Architecture Comparison: Throughput", fontsize=14, fontweight='bold')
    plt.ylabel("Total Events Processed")
    
    plt.figtext(0.5, -0.05, f"Stream Speed: {events_count/duration:.2f} events/sec", wrap=True, horizontalalignment='center', fontsize=10)
    
    plt.tight_layout()
    plt.savefig(filename, dpi=300)
    plt.close()
    print(f"[INFO] Kafka Stream Plot saved to {os.path.abspath(filename)}")

if __name__ == "__main__":
    # Test script standalone (requires preceding steps)
    from data_ingestion import fetch_match_data, process_and_flatten_data
    from qos_engine import calculate_qos_index, apply_hardware_fatigue_simulation, detect_critical_nodes
    
    # Run the Pipeline
    MATCH_ID = 303596 
    raw = fetch_match_data(MATCH_ID)
    processed = process_and_flatten_data(raw)
    qos_data = calculate_qos_index(processed)
    fatigue_data = apply_hardware_fatigue_simulation(qos_data)
    final_data, faults = detect_critical_nodes(fatigue_data)
    
    # reporting
    generate_incident_log(faults)
    plot_degradation_timeline(final_data, team_name="Barcelona")
